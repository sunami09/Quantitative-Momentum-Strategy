import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def to_excel(df):
    filename = 'datasource.xlsx'
    # Write the DataFrame to an Excel file
    df.to_excel(filename, index=False)

    # Load the workbook
    book = load_workbook(filename)

    # Get the worksheet you're working with
    sheet = book.active  # This gets the active worksheet

    # Iterate over the columns and set the width
    for i, column in enumerate(sheet.columns, start=1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    # Save your changes
    book.save(filename)
